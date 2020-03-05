# -*- coding: utf-8 -*-
import os
from docx  import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

info = {"委托人":[],"委托鉴定事项":[],"受理日期":[],"鉴定材料":[],"鉴定日期":[],"鉴定地点":[],"被鉴定人":[],"鉴定意见":[],"文件路径":""}
def gen_docx_list(path,ret):
    filelist = os.listdir(path)
    for filename in filelist:
        de_path = os.path.join(path, filename)
        if os.path.isfile(de_path):
            if de_path.endswith(".docx"):  # Specify to find the docx file.
                ret.append(de_path)
        else:
            gen_docx_list(de_path,ret)
    return ret
    

def analy_docx(file_path):
    index = [0 for i in range(len(info))]
    
    document=Document(file_path)
    lines = []
    for paragraph in document.paragraphs:
        lines.append(paragraph.text)
   
    info["文件路径"] = file_path
    for line in lines:
        if line.startswith('委托人'):
            info["委托人"] = line[4::]
        if line.startswith('委托鉴定事项'):
            info["委托鉴定事项"] = line[7::]
        if line.startswith("受理日期"):
            info["受理日期"] = line[5::]
        if line.startswith("鉴定材料"):
            info["鉴定材料"] = line[5::] 
        if line.startswith("鉴定日期"):
            info["鉴定日期"] = line[5::]
        if line.startswith("鉴定地点"):
            info["鉴定地点"] = line[5::]
        if line.startswith("被鉴定人"):
            info["被鉴定人"] = line[5::]
        if "损伤程度评为" in line:
            info["鉴定意见"] = line
    return info  

#设置excel文件格式
def foamrt_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]
    for column in ws.columns:
        for cell in column:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    #设置行高度
    for i in range(1,ws.max_row + 1):
        ws.row_dimensions[i].height = 45
    #设置列宽度
    for i in range(1,ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    wb.save(excel_path)    


if __name__ == '__main__':
    path=os.getcwd()
    doclist=gen_docx_list(path,[])
    workbook = Workbook()
    booksheet = workbook.active
    title = [i for i in info.keys()]
    booksheet.append(title)
    
    for file in doclist:
        info=analy_docx(file)
        info_to_excel = [i for i in info.values()]
        booksheet.append(info_to_excel)
        workbook.save("reslut.xlsx")
        
    workbook.close()
    
    #将生成的excel文件格式进行设置
    foamrt_excel("reslut.xlsx")