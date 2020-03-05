from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
 
wb = load_workbook("demo2.xlsx")
ws = wb[wb.sheetnames[0]]
for column in ws.columns:
    for cell in column:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#设置行高度
for i in range(1,ws.max_row + 1):
    ws.row_dimensions[i].height = 45
#设置列宽度
for i in range(1,ws.max_column + 1):
    ws.column_dimensions[get_column_letter(i)].width = 15
wb.save('demo2.xlsx')